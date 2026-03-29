"""
Import vocabulary Excel (Topics, Vocabulary, VocabExamples) into SQL Server.
Target schema: data/db/prod/VocabifySchema.sql (Topics.IsActive / soft-delete columns as defined there).

Default run (no --excel): imports all 5 Excel files (E, M, I, UI, A) under data/excel/.

Single file: pass --excel and optional --level-code.

Connection (choose one):
  - Env VOCABIFY_SQL_CONNECTION_STRING (full ODBC connection string), or
  - VOCABIFY_SQL_SERVER, VOCABIFY_SQL_DATABASE, and optionally VOCABIFY_SQL_USER + VOCABIFY_SQL_PASSWORD
    (omit user/password for Windows auth).

Example (PowerShell):
  $env:VOCABIFY_SQL_SERVER="PT"
  $env:VOCABIFY_SQL_DATABASE="Vocabify"

  # All 5 levels (default)
  python import_vocab_excel_to_sqlserver.py

  # One file only
  python import_vocab_excel_to_sqlserver.py --excel ..\\..\\excel\\vocab_easy_10topics_10words.xlsx --level-code E

  # Only M, I, UI, A (when Easy is already imported)
  python import_vocab_excel_to_sqlserver.py --import-remaining-levels
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from pathlib import Path

import pyodbc
from openpyxl import load_workbook

# Script lives under data/db/import/ — Excel files are under data/excel/
# parents[0]=import dir, [1]=db, [2]=data
_DATA_ROOT = Path(__file__).resolve().parents[2]
EXCEL_DIR = _DATA_ROOT / "excel"

# Excel filename → Levels.Code (see VocabifySchema.sql seed)
ALL_LEVEL_FILES: list[tuple[str, str]] = [
    ("vocab_easy_10topics_10words.xlsx", "E"),
    ("vocab_medium_10topics_10words.xlsx", "M"),
    ("vocab_intermediate_10topics_10words.xlsx", "I"),
    ("vocab_upper_intermediate_10topics_10words.xlsx", "UI"),
    ("vocab_advanced_10topics_10words.xlsx", "A"),
]

# Subset: all except Easy (same codes as ALL_LEVEL_FILES[1:])
REMAINING_LEVEL_FILES: list[tuple[str, str]] = ALL_LEVEL_FILES[1:]


def build_connection_string() -> str:
    full = os.environ.get("VOCABIFY_SQL_CONNECTION_STRING", "").strip()
    if full:
        return full
    server = os.environ.get("VOCABIFY_SQL_SERVER", "PT").strip()
    database = os.environ.get("VOCABIFY_SQL_DATABASE", "Vocabify").strip()
    user = os.environ.get("VOCABIFY_SQL_USER", "").strip()
    password = os.environ.get("VOCABIFY_SQL_PASSWORD", "").strip()
    if user:
        return (
            "Driver={ODBC Driver 17 for SQL Server};"
            f"Server={server};Database={database};Uid={user};Pwd={password};"
            "TrustServerCertificate=yes;"
        )
    return (
        "Driver={ODBC Driver 17 for SQL Server};"
        f"Server={server};Database={database};Trusted_Connection=yes;"
        "TrustServerCertificate=yes;"
    )


def slugify_topic(name: str) -> str:
    s = (name or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    s = re.sub(r"-+", "-", s).strip("-")
    return s or "topic"


def trunc(s: str | None, max_len: int | None) -> str | None:
    if s is None:
        return None
    t = str(s).strip()
    if max_len is None or len(t) <= max_len:
        return t
    return t[:max_len]


def parse_examples_cell(raw) -> list[dict]:
    if raw is None or raw == "":
        return []
    if isinstance(raw, list):
        return [x for x in raw if isinstance(x, dict)]
    text = str(raw).strip()
    if not text:
        return []
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        return []
    if not isinstance(data, list):
        return []
    out = []
    for item in data:
        if isinstance(item, dict):
            en = (item.get("en") or item.get("En") or "").strip()
            vi = (item.get("vi") or item.get("Vi") or "").strip()
            if en:
                out.append({"en": en, "vi": vi})
    return out


def read_excel_rows(path: Path) -> tuple[list[str], list[dict]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        wb.close()
        return [], []
    headers: list[str | None] = []
    for c in header_row:
        if c is None or (isinstance(c, str) and not str(c).strip()):
            headers.append(None)
        else:
            headers.append(str(c).strip())
    name_to_idx = {h: i for i, h in enumerate(headers) if h}

    required = {"word", "topic"}
    missing = [k for k in required if k not in name_to_idx]
    if missing:
        wb.close()
        raise ValueError(f"Excel missing columns: {missing}. Found: {list(name_to_idx.keys())}")

    data_rows: list[dict] = []
    for row in rows_iter:
        if not row or all(v is None or (isinstance(v, str) and not str(v).strip()) for v in row):
            continue
        rec: dict = {}
        for name, idx in name_to_idx.items():
            rec[name] = row[idx] if idx < len(row) else None
        w = rec.get("word")
        if w is None or (isinstance(w, str) and not str(w).strip()):
            continue
        data_rows.append(rec)

    wb.close()
    return [h for h in headers if h], data_rows


def resolve_level_id(cur, level_code: str) -> int:
    cur.execute(
        """
        SELECT TOP 1 Id FROM dbo.Levels
        WHERE LOWER(LTRIM(RTRIM(Code))) = LOWER(LTRIM(RTRIM(?)))
           OR LOWER(LTRIM(RTRIM(Name))) = LOWER(LTRIM(RTRIM(?)))
        ORDER BY Id
        """,
        (level_code, level_code),
    )
    row = cur.fetchone()
    if row:
        return int(row[0])
    cur.execute("SELECT Code, Name FROM dbo.Levels ORDER BY SortOrder")
    codes = cur.fetchall()
    hint = "; ".join(f"{r[0]}={r[1]}" for r in codes) if codes else "(no rows)"
    raise RuntimeError(
        f"Level not found for code/name {level_code!r}. Existing levels: {hint}"
    )


def next_topic_sort_order(cur, level_id: int) -> int:
    cur.execute(
        "SELECT COALESCE(MAX(SortOrder), 0) + 1 FROM dbo.Topics WHERE LevelId = ? AND IsDeleted = 0",
        (level_id,),
    )
    return int(cur.fetchone()[0])


def get_or_create_topic(
    cur, level_id: int, topic_name: str, slug: str, sort_order: int
) -> int:
    cur.execute(
        """
        SELECT Id FROM dbo.Topics
        WHERE LevelId = ? AND Slug = ? AND IsDeleted = 0
        """,
        (level_id, slug),
    )
    row = cur.fetchone()
    if row:
        return int(row[0])

    display_name = topic_name.strip() if topic_name else slug
    display_name = trunc(display_name, 200) or slug
    cur.execute(
        """
        INSERT INTO dbo.Topics (LevelId, Slug, Name, Description, SortOrder, IsActive, IsDeleted)
        OUTPUT INSERTED.Id
        VALUES (?, ?, ?, NULL, ?, 1, 0)
        """,
        (level_id, slug, display_name, sort_order),
    )
    new_id = cur.fetchone()[0]
    return int(new_id)


def vocab_exists(cur, topic_id: int, word: str) -> bool:
    cur.execute(
        """
        SELECT 1 FROM dbo.Vocabulary
        WHERE TopicId = ? AND Word = ? AND IsDeleted = 0
        """,
        (topic_id, word),
    )
    return cur.fetchone() is not None


def example_exists(cur, vocabulary_id: int, example_en: str) -> bool:
    cur.execute(
        """
        SELECT 1 FROM dbo.VocabExamples
        WHERE VocabularyId = ? AND ExampleEn = ? AND IsDeleted = 0
        """,
        (vocabulary_id, example_en),
    )
    return cur.fetchone() is not None


def run_import(
    cur, path: Path, level_code: str
) -> tuple[int, int, int, int, int]:
    """
    Insert one workbook for one level. Does not commit.
    Returns: level_id, inserted_vocab, skipped_vocab, inserted_examples, row_count
    """
    _, data_rows = read_excel_rows(path)
    if not data_rows:
        raise ValueError(f"No data rows in {path}")

    topic_order: list[str] = []
    seen: set[str] = set()
    for r in data_rows:
        t = (r.get("topic") or "").strip()
        if not t:
            continue
        if t not in seen:
            seen.add(t)
            topic_order.append(t)

    level_id = resolve_level_id(cur, level_code)

    slug_by_topic: dict[str, str] = {t: slugify_topic(t) for t in topic_order}
    topic_id_by_name: dict[str, int] = {}
    sort_cursor = next_topic_sort_order(cur, level_id)

    for t in topic_order:
        slug = slug_by_topic[t]
        tid = get_or_create_topic(cur, level_id, t, slug, sort_cursor)
        topic_id_by_name[t] = tid
        cur.execute(
            "SELECT SortOrder FROM dbo.Topics WHERE Id = ?",
            (tid,),
        )
        so = int(cur.fetchone()[0])
        if so >= sort_cursor:
            sort_cursor = so + 1

    inserted_vocab = 0
    skipped_vocab = 0
    inserted_examples = 0

    for r in data_rows:
        topic_name = (r.get("topic") or "").strip()
        if not topic_name or topic_name not in topic_id_by_name:
            continue
        topic_id = topic_id_by_name[topic_name]
        word_raw = r.get("word")
        word = trunc(str(word_raw).strip(), 100)
        if not word:
            continue
        if vocab_exists(cur, topic_id, word):
            skipped_vocab += 1
            continue

        phonetic = trunc(r.get("phonetic"), 100)
        audio = trunc(r.get("audio"), 500)
        pos = trunc(r.get("part_of_speech"), 50)
        def_en = trunc(r.get("definition_en"), None)
        def_vi = trunc(r.get("definition_vi"), None)
        meaning_vi = trunc(r.get("meaning_vi"), 100)

        cur.execute(
            """
            INSERT INTO dbo.Vocabulary (
                TopicId, Word, Phonetic, AudioUrl, PartOfSpeech,
                DefinitionEn, DefinitionVi, MeaningVi, IsDeleted
            )
            OUTPUT INSERTED.Id
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 0)
            """,
            (
                topic_id,
                word,
                phonetic,
                audio,
                pos,
                def_en,
                def_vi,
                meaning_vi,
            ),
        )
        vid = int(cur.fetchone()[0])
        inserted_vocab += 1

        examples = parse_examples_cell(r.get("examples"))
        for ex in examples:
            ex_en = trunc(ex.get("en"), 1000)
            if not ex_en:
                continue
            if example_exists(cur, vid, ex_en):
                continue
            ex_vi = trunc(ex.get("vi"), 1000)
            cur.execute(
                """
                INSERT INTO dbo.VocabExamples (VocabularyId, ExampleEn, ExampleVi, IsDeleted)
                VALUES (?, ?, ?, 0)
                """,
                (vid, ex_en, ex_vi),
            )
            inserted_examples += 1

    return level_id, inserted_vocab, skipped_vocab, inserted_examples, len(data_rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="Import vocab Excel to SQL Server")
    parser.add_argument(
        "--import-remaining-levels",
        action="store_true",
        help="Import only M, I, UI, A (4 files). Mutually exclusive with --excel.",
    )
    parser.add_argument(
        "--excel",
        type=Path,
        default=None,
        help="Import a single .xlsx. If omitted, imports all 5 level files under data/excel/.",
    )
    parser.add_argument(
        "--level-code",
        default="E",
        help="Levels.Code for single-file import only (default E)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Read Excel and print counts only; no DB writes",
    )
    args = parser.parse_args()

    if args.import_remaining_levels and args.excel is not None:
        print("Do not combine --excel with --import-remaining-levels.", file=sys.stderr)
        return 1

    if args.import_remaining_levels:
        job_list = REMAINING_LEVEL_FILES
    elif args.excel is not None:
        job_list = None
    else:
        job_list = ALL_LEVEL_FILES

    if job_list is not None:
        jobs: list[tuple[Path, str]] = []
        for fname, code in job_list:
            p = EXCEL_DIR / fname
            if not p.is_file():
                print(f"Missing file: {p}", file=sys.stderr)
                return 1
            jobs.append((p, code))

        if args.dry_run:
            for path, code in jobs:
                _, data_rows = read_excel_rows(path)
                topic_order: list[str] = []
                seen: set[str] = set()
                for r in data_rows:
                    t = (r.get("topic") or "").strip()
                    if not t:
                        continue
                    if t not in seen:
                        seen.add(t)
                        topic_order.append(t)
                print(f"[dry-run] {code} | {path.name} | rows={len(data_rows)} topics={len(topic_order)}")
            return 0

        conn_str = build_connection_string()
        try:
            conn = pyodbc.connect(conn_str, autocommit=False)
        except pyodbc.Error as e:
            print(f"DB connection failed: {e}", file=sys.stderr)
            print(
                "Set VOCABIFY_SQL_CONNECTION_STRING or VOCABIFY_SQL_SERVER + VOCABIFY_SQL_DATABASE.",
                file=sys.stderr,
            )
            return 1

        cur = conn.cursor()
        try:
            for path, code in jobs:
                print(f"Importing level {code}: {path.name} ...")
                (
                    level_id,
                    ins_v,
                    skip_v,
                    ins_e,
                    nrows,
                ) = run_import(cur, path, code)
                conn.commit()
                print(
                    f"  Done LevelId={level_id} rows={nrows} | vocab +{ins_v} skipped={skip_v} | examples +{ins_e}"
                )
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()
        return 0

    path: Path = args.excel
    if not path.is_file():
        print(f"File not found: {path}", file=sys.stderr)
        return 1

    _, data_rows = read_excel_rows(path)
    if not data_rows:
        print("No data rows in Excel.", file=sys.stderr)
        return 1

    topic_order_single: list[str] = []
    seen_s: set[str] = set()
    for r in data_rows:
        t = (r.get("topic") or "").strip()
        if not t:
            continue
        if t not in seen_s:
            seen_s.add(t)
            topic_order_single.append(t)

    print(f"Excel: {path}")
    print(f"Rows: {len(data_rows)} | Distinct topics: {len(topic_order_single)}")

    if args.dry_run:
        return 0

    conn_str = build_connection_string()
    try:
        conn = pyodbc.connect(conn_str, autocommit=False)
    except pyodbc.Error as e:
        print(f"DB connection failed: {e}", file=sys.stderr)
        print(
            "Set VOCABIFY_SQL_CONNECTION_STRING or VOCABIFY_SQL_SERVER + VOCABIFY_SQL_DATABASE.",
            file=sys.stderr,
        )
        return 1

    cur = conn.cursor()
    try:
        level_id, inserted_vocab, skipped_vocab, inserted_examples, _ = run_import(
            cur, path, args.level_code
        )
        conn.commit()
        print(
            f"Done. LevelId={level_id} | Vocabulary inserted={inserted_vocab} skipped={skipped_vocab} | Examples inserted={inserted_examples}"
        )
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
